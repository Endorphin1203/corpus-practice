#!/usr/bin/env python3
"""直接导入 Original 目录下的两个 Excel 到 MySQL 数据库。"""

import openpyxl
import pymysql
import os
import re

DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': os.environ.get('DB_PASSWORD', 'root'),
    'database': 'corpus_practice',
    'charset': 'utf8mb4',
}

BASE = '/Users/a1-6/Documents/ClaudeCode/2026-07-16/Original'

# ============================================================
# 描写续写.xlsx — 5 个 Sheet
# ============================================================
def import_miaoxie(conn):
    path = os.path.join(BASE, '（自整）描写续写.xlsx')
    wb = openpyxl.load_workbook(path)
    total = 0

    # --- 优秀表达: col 0=中文, col 1=英文, col 2=备注, from row 2 ---
    ws = wb['优秀表达']
    for row in ws.iter_rows(min_row=2, values_only=True):
        ch, en, note = row[0], row[1], row[2] if len(row) > 2 else None
        if not ch or not en:
            continue
        insert(conn, str(ch).strip(), str(en).strip(), str(note).strip() if note else None,
               '描写续写', '优秀表达')
        total += 1
    print(f'  优秀表达: {total} 条')

    # --- 动作: col 0=英文, col 1=中文, from row 2, skip emoji header rows ---
    ws = wb['动作']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        en, ch = row[0], row[1] if len(row) > 1 else None
        if not en or not ch:
            continue
        en_str = str(en).strip()
        # 跳过子分类标题行（如 "走🦶"）
        if len(en_str) <= 5 and any(ord(c) > 127 for c in en_str if not c.isascii()):
            continue
        insert(conn, str(ch).strip(), en_str, None, '描写续写', '动作')
        count += 1
    print(f'  动作: {count} 条')
    total += count

    # --- 情绪: col 0=英文, col 1=中文, col 2=备注, from row 3 (skip row 2=恐惧😱) ---
    ws = wb['情绪']
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        en, ch, note = row[0], row[1], row[2] if len(row) > 2 else None
        if not ch or not en:
            continue
        ch_str = str(ch).strip()
        en_str = str(en).strip()
        # 跳过子分类标题（如 "恐惧😱" 的行，特征是中文为 None 或空）
        if not ch_str:
            continue
        insert(conn, ch_str, en_str, str(note).strip() if note else None,
               '描写续写', '情绪')
        count += 1
    print(f'  情绪: {count} 条')
    total += count

    # --- 环境: col 0=英文, col 1=中文, from row 2 ---
    ws = wb['环境']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        en, ch = row[0], row[1]
        if not ch or not en:
            continue
        insert(conn, str(ch).strip(), str(en).strip(), None, '描写续写', '环境')
        count += 1
    print(f'  环境: {count} 条')
    total += count

    # --- 外貌: col 0=英文, col 1=中文, from row 2 ---
    ws = wb['外貌']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        en, ch = row[0], row[1]
        if not ch or not en:
            continue
        insert(conn, str(ch).strip(), str(en).strip(), None, '描写续写', '外貌')
        count += 1
    print(f'  外貌: {count} 条')
    total += count

    print(f'  描写续写合计: {total} 条')


# ============================================================
# 议论文.xlsx — 6 个 Sheet
# ============================================================
def import_yilunwen(conn):
    path = os.path.join(BASE, '（自整）议论文.xlsx')
    wb = openpyxl.load_workbook(path)
    total = 0

    # --- 好句段落: col 0=中文, col 1=英文, col 2=备注, from row 2 ---
    ws = wb['好句段落']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        ch, en = row[0], row[1]
        note = row[2] if len(row) > 2 else None
        if not ch or not en:
            continue
        insert(conn, str(ch).strip(), str(en).strip(), str(note).strip() if note else None,
               '议论文', '好句段落')
        count += 1
    print(f'  好句段落: {count} 条')
    total += count

    # --- 开头: col 0=英文句型, col 1=中文说明, col 2=例句, from row 2 ---
    # 跳过子分类标题行（中文说明为 None）
    ws = wb['开头']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        en, ch, example = row[0], row[1], row[2] if len(row) > 2 else None
        if not en or not ch:
            # 子分类标题行（如第 2 行 "正反两点论"）
            continue
        note = str(example).strip() if example else None
        ch_str = str(ch).strip()
        en_str = str(en).strip()
        insert(conn, ch_str, en_str, note, '议论文', '开头')
        count += 1
    print(f'  开头: {count} 条')
    total += count

    # --- 主体: col 0=英文句型, col 1=中文说明, col 2=例句, from row 2 ---
    ws = wb['主体']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        en, ch, example = row[0], row[1], row[2] if len(row) > 2 else None
        if not en or not ch:
            continue
        insert(conn, str(ch).strip(), str(en).strip(),
               str(example).strip() if example else None, '议论文', '主体')
        count += 1
    print(f'  主体: {count} 条')
    total += count

    # --- 观点库: col 0=英文, col 1=中文, col 2=例句, from row 2 ---
    ws = wb['观点库']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        en, ch, example = row[0], row[1], row[2] if len(row) > 2 else None
        if not ch or not en:
            continue
        insert(conn, str(ch).strip(), str(en).strip(),
               str(example).strip() if example else None, '议论文', '观点库')
        count += 1
    print(f'  观点库: {count} 条')
    total += count

    # --- 高级表达: col 0=中文模式说明, col 1=英文, col 2=例句, from row 2 ---
    ws = wb['高级表达']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        ch, en, example = row[0], row[1], row[2] if len(row) > 2 else None
        if not ch or not en:
            continue
        insert(conn, str(ch).strip(), str(en).strip(),
               str(example).strip() if example else None, '议论文', '高级表达')
        count += 1
    print(f'  高级表达: {count} 条')
    total += count

    # --- 主题词: col 0=中文, col 1=英文, from row 2 ---
    ws = wb['主题词']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        ch, en = row[0], row[1]
        if not ch or not en:
            continue
        insert(conn, str(ch).strip(), str(en).strip(), None, '议论文', '主题词')
        count += 1
    print(f'  主题词: {count} 条')
    total += count

    print(f'  议论文合计: {total} 条')


def insert(conn, chinese, english, notes, category, subcategory):
    """插入一条语料，按 chinese 去重。"""
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM corpus WHERE chinese = %s", (chinese,))
        if cur.fetchone():
            return  # 已存在，跳过
        cur.execute(
            "INSERT INTO corpus (chinese, english, notes, category, subcategory) VALUES (%s, %s, %s, %s, %s)",
            (chinese, english, notes, category, subcategory))
    conn.commit()


def main():
    conn = pymysql.connect(**DB_CONFIG)
    print(f'连接数据库成功: {DB_CONFIG["database"]}')

    # 清空旧数据
    with conn.cursor() as cur:
        cur.execute("DELETE FROM practice_answer")
        cur.execute("DELETE FROM practice_session")
        cur.execute("DELETE FROM exercise_cache")
        cur.execute("DELETE FROM corpus")
    conn.commit()
    print('已清空旧数据\n')

    print('--- 导入描写续写 ---')
    import_miaoxie(conn)

    print('\n--- 导入议论文 ---')
    import_yilunwen(conn)

    # 统计
    with conn.cursor() as cur:
        cur.execute("SELECT category, subcategory, COUNT(*) FROM corpus GROUP BY category, subcategory ORDER BY category, subcategory")
        print('\n=== 导入结果 ===')
        for row in cur.fetchall():
            print(f'  {row[0]} / {row[1]}: {row[2]} 条')
        cur.execute("SELECT COUNT(*) FROM corpus")
        print(f'\n总计: {cur.fetchone()[0]} 条语料')

    conn.close()
    print('\n导入完成！')


if __name__ == '__main__':
    main()
